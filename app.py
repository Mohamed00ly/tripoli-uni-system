import os 
from functools import wraps
from flask import (Flask, render_template, redirect, url_for, flash,
                   request, abort, send_file, send_from_directory)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
from sqlalchemy import func, extract
import io, csv, os, uuid

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import Config
from models import (db, User, PreRegisteredStudent, Department, Course,
                    Enrollment, AuditLog, GradeFormula, CourseAssignment,
                    ClassSchedule, Classroom, Document, SystemSettings)

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة'
login_manager.login_message_category = 'warning'


@app.context_processor
def inject_system_settings():
    """Inject system_logo_url into every template context."""
    try:
        settings  = SystemSettings.query.first()
        logo_path = settings.logo_path if (settings and settings.logo_path) else 'img/logo.png'
    except Exception:
        logo_path = 'img/logo.png'
    return {'system_logo_url': url_for('static', filename=logo_path)}

ALLOWED_IMAGES = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
ALLOWED_DOCS   = {'pdf', 'jpg', 'jpeg', 'png'}


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ── Decorators ────────────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*a, **kw)
    return d


def lecturer_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not current_user.is_authenticated or current_user.is_student:
            abort(403)
        return f(*a, **kw)
    return d


def full_access_required(f):
    """Only admins with full_access (or NULL legacy) permissions can proceed."""
    @wraps(f)
    def d(*a, **kw):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        if not current_user.has_full_access:
            flash('هذه العملية تتطلب صلاحية "الوصول الكامل"', 'error')
            return redirect(url_for('admin_dashboard'))
        return f(*a, **kw)
    return d


# ── Helpers ───────────────────────────────────────────────────────────────────

def validate_email_domain(email):
    e = email.lower()
    return e.endswith('@gmail.com') or e.endswith('@yahoo.com')


def log_action(action, entity_type, entity_id=None, old_val=None, new_val=None):
    try:
        db.session.add(AuditLog(
            user_id=current_user.id if current_user.is_authenticated else None,
            action=action, entity_type=entity_type, entity_id=entity_id,
            old_value=str(old_val) if old_val is not None else None,
            new_value=str(new_val) if new_val is not None else None,
            ip_address=request.remote_addr,
        ))
    except Exception:
        pass


def compute_grade(enrollment):
    if enrollment.classwork_grade is None or enrollment.final_grade is None:
        return None
    f = GradeFormula.query.filter_by(course_id=enrollment.course_id).first()
    cw, fw = (f.classwork_weight, f.final_weight) if f else (40, 60)
    return round((enrollment.classwork_grade * cw + enrollment.final_grade * fw) / 100, 2)


def save_upload(file, subfolder, allowed_ext):
    if not file or not file.filename:
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in allowed_ext:
        return None
    unique_name = f"{uuid.uuid4().hex}.{ext}"
    save_dir = os.path.join(app.config['UPLOAD_FOLDER'], subfolder)
    os.makedirs(save_dir, exist_ok=True)
    file.save(os.path.join(save_dir, unique_name))
    return f"{subfolder}/{unique_name}"


def _render_transcript(student):
    enrollments = Enrollment.query.filter_by(user_id=student.id).all()
    total_credits = total_points = 0.0
    for e in enrollments:
        if e.grade is not None:
            c = e.course.credits
            total_credits += c
            total_points  += (e.grade / 100) * 4.0 * c
    gpa = round(total_points / total_credits, 2) if total_credits else 0.0
    return render_template('student/transcript.html',
                           student=student, enrollments=enrollments, gpa=gpa)


def _save_docs_for(entity_key_field, entity_id, doc_types=('birth_certificate', 'qualification')):
    """Save documents; entity_key_field is 'user_id' or 'pre_reg_id'."""
    for doc_type in doc_types:
        file = request.files.get(doc_type)
        path = save_upload(file, 'documents', ALLOWED_DOCS)
        if path:
            filt = {entity_key_field: entity_id, 'document_type': doc_type}
            old = Document.query.filter_by(**filt).first()
            if old:
                old.filename = path
                old.original_name = secure_filename(file.filename)
            else:
                doc = Document(document_type=doc_type, filename=path,
                               original_name=secure_filename(file.filename))
                setattr(doc, entity_key_field, entity_id)
                db.session.add(doc)


# ── Static uploads ────────────────────────────────────────────────────────────

@app.route('/uploads/<path:filename>')
@login_required
def serve_upload(filename):
    if '..' in filename or filename.startswith('/'):
        abort(400)
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# ── Index ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard'))
    if current_user.is_lecturer:
        return redirect(url_for('lecturer_dashboard'))
    return redirect(url_for('dashboard'))


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            log_action('تسجيل دخول', 'user', user.id)
            db.session.commit()
            return redirect(request.args.get('next') or url_for('index'))
        flash('البريد الإلكتروني أو كلمة المرور غير صحيحة', 'error')
    return render_template('auth/login.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        full_name  = request.form.get('full_name', '').strip()
        email      = request.form.get('email', '').strip().lower()
        password   = request.form.get('password', '')
        confirm    = request.form.get('confirm_password', '')

        if not validate_email_domain(email):
            flash('يُقبل فقط البريد من @gmail.com أو @yahoo.com', 'error')
            return render_template('auth/signup.html')
        pre = PreRegisteredStudent.query.filter_by(
            student_id=student_id, full_name=full_name).first()
        if not pre:
            flash('رقم القيد أو الاسم غير مدرج، يرجى مراجعة الإدارة', 'error')
            return render_template('auth/signup.html')
        if User.query.filter_by(student_id=student_id).first():
            flash('رقم القيد مسجل بحساب موجود مسبقاً', 'error')
            return render_template('auth/signup.html')
        if User.query.filter_by(email=email).first():
            flash('البريد الإلكتروني مستخدم مسبقاً', 'error')
            return render_template('auth/signup.html')
        if password != confirm:
            flash('كلمة المرور وتأكيدها غير متطابقتين', 'error')
            return render_template('auth/signup.html')
        if len(password) < 8:
            flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
            return render_template('auth/signup.html')

        user = User(
            student_id=student_id, full_name=full_name, email=email,
            password_hash=generate_password_hash(password),
            role='student', status='pending', financial_status='unpaid',
            department_id=pre.department_id,
        )
        db.session.add(user)
        db.session.flush()
        log_action('إنشاء حساب', 'user', user.id, new_val=email)
        db.session.commit()
        flash('تم إنشاء حسابك. ستتم مراجعته من قبل الإدارة.', 'success')
        return redirect(url_for('login'))
    return render_template('auth/signup.html')


@app.route('/logout')
@login_required
def logout():
    log_action('تسجيل خروج', 'user', current_user.id)
    db.session.commit()
    logout_user()
    return redirect(url_for('login'))


# ── Student ───────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    if not current_user.is_student:
        return redirect(url_for('index'))
    enrollments = Enrollment.query.filter_by(user_id=current_user.id).all()
    return render_template('student/dashboard.html', enrollments=enrollments)


def _is_first_semester(user_id):
    """True when the student has no graded enrollment records (never completed a semester)."""
    return not Enrollment.query.filter(
        Enrollment.user_id == user_id,
        Enrollment.grade.isnot(None)
    ).first()


def _active_enrollment_count(user_id, semester, year):
    """Count of non-admin-dropped enrollments for the given semester/year."""
    return Enrollment.query.filter(
        Enrollment.user_id == user_id,
        Enrollment.semester == semester,
        Enrollment.year == year,
        Enrollment.dropped_by_admin == False  # noqa: E712
    ).count()


@app.route('/enrollment')
@login_required
def enrollment():
    if not current_user.is_student:
        return redirect(url_for('index'))
    if current_user.status != 'approved':
        flash('حسابك لم يُوافق عليه بعد.', 'warning')
        return redirect(url_for('dashboard'))
    if current_user.financial_status == 'unpaid':
        return render_template('student/enrollment_locked.html')

    # Guard: student must be assigned to a department before they can enroll
    if not current_user.department_id:
        flash('لم يتم تعيين قسم أكاديمي لحسابك بعد. '
              'يرجى التواصل مع مسجّل الكلية لتحديث بياناتك.', 'warning')
        return render_template('student/enrollment.html', courses_data=[],
                               is_first_sem=True, current_count=0)

    is_first_sem  = _is_first_semester(current_user.id)
    current_count = _active_enrollment_count(current_user.id, '1', 2024)

    all_current = Enrollment.query.filter_by(
        user_id=current_user.id, semester='1', year=2024).all()
    enrolled_this     = {e.course_id for e in all_current if not e.dropped_by_admin}
    admin_dropped_ids = {e.course_id for e in all_current if e.dropped_by_admin}

    all_enrollments = Enrollment.query.filter_by(user_id=current_user.id).all()
    passed_ids = {
        e.course_id for e in all_enrollments
        if e.grade is not None and e.grade >= 50 and not e.dropped_by_admin
    }
    failed_ids = {
        e.course_id for e in all_enrollments
        if e.grade is not None and e.grade < 50 and not e.dropped_by_admin
    }

    # Strict department filter — only courses belonging to the student's own department
    dept_courses = Course.query.filter_by(
        department_id=current_user.department_id
    ).order_by(Course.code).all()

    courses_data = []
    for course in dept_courses:
        can = True; reason = ''
        is_failed = course.id in failed_ids

        if is_first_sem:
            # First-semester students: no prerequisite checks; enforce 6-course cap
            if current_count >= 6 and course.id not in enrolled_this:
                can = False
                reason = 'تجاوزت الحد الأقصى للفصل الأول (6 مواد)'
        else:
            # Second+ semester: enforce prerequisites (retaking a failed course is exempt)
            if course.prerequisite_id and not is_failed \
                    and course.prerequisite_id not in passed_ids:
                can = False
                prereq = Course.query.get(course.prerequisite_id)
                reason = f'يتطلب اجتياز: {prereq.name_ar}' if prereq else 'متطلب سابق'

        courses_data.append({
            'course':          course,
            'can_enroll':      can,
            'reason':          reason,
            'is_enrolled':     course.id in enrolled_this,
            'is_admin_dropped': course.id in admin_dropped_ids,
            'is_failed':       is_failed,
        })

    return render_template('student/enrollment.html',
                           courses_data=courses_data,
                           is_first_sem=is_first_sem,
                           current_count=current_count)


@app.route('/enroll/<int:course_id>', methods=['POST'])
@login_required
def enroll_course(course_id):
    if not current_user.is_student:
        abort(403)
    if current_user.financial_status == 'unpaid':
        flash('الرسوم الجامعية غير مدفوعة', 'error')
        return redirect(url_for('enrollment'))

    course       = Course.query.get_or_404(course_id)
    is_first_sem = _is_first_semester(current_user.id)

    # Check if student has a prior failed attempt (retake grace exemption)
    past_failed = Enrollment.query.filter(
        Enrollment.user_id  == current_user.id,
        Enrollment.course_id == course_id,
        Enrollment.grade.isnot(None),
        Enrollment.grade     < 50
    ).first()

    if is_first_sem:
        # First-semester cap: max 6 courses, no prerequisite checks
        current_count = _active_enrollment_count(current_user.id, '1', 2024)
        if current_count >= 6:
            flash('تجاوزت الحد الأقصى للمواد في الفصل الأول (6 مواد)', 'error')
            return redirect(url_for('enrollment'))
    else:
        # Second+ semester: enforce prerequisites unless retaking a failed course
        if course.prerequisite_id and not past_failed:
            passed = Enrollment.query.filter(
                Enrollment.user_id  == current_user.id,
                Enrollment.course_id == course.prerequisite_id,
                Enrollment.grade    >= 50
            ).first()
            if not passed:
                flash('لم تجتز المتطلب السابق', 'error')
                return redirect(url_for('enrollment'))

    existing = Enrollment.query.filter_by(
        user_id=current_user.id, course_id=course_id, semester='1', year=2024).first()
    if existing:
        if existing.dropped_by_admin:
            existing.dropped_by_admin = False
            existing.status = 'enrolled'
            log_action('إعادة تسجيل بعد إسقاط إداري', 'enrollment', course_id, new_val=course.code)
            db.session.commit()
            flash(f'تم التسجيل في {course.name_ar}', 'success')
        else:
            flash('مسجل في هذه المادة مسبقاً', 'warning')
        return redirect(url_for('enrollment'))

    db.session.add(Enrollment(
        user_id=current_user.id, course_id=course_id, semester='1', year=2024))
    log_action('تسجيل مادة', 'enrollment', course_id, new_val=course.code)
    db.session.commit()
    flash(f'تم التسجيل في {course.name_ar}', 'success')
    return redirect(url_for('enrollment'))


@app.route('/unenroll/<int:course_id>', methods=['POST'])
@login_required
def unenroll_course(course_id):
    if current_user.financial_status == 'paid':
        flash('عذراً، لا يمكن إسقاط المادة بعد إتمام عملية دفع التكاليف.', 'error')
        return redirect(url_for('enrollment'))
    rec = Enrollment.query.filter_by(
        user_id=current_user.id, course_id=course_id, semester='1', year=2024
    ).first_or_404()
    log_action('إلغاء تسجيل', 'enrollment', course_id)
    db.session.delete(rec)
    db.session.commit()
    flash('تم إلغاء التسجيل', 'success')
    return redirect(url_for('enrollment'))


@app.route('/transcript')
@login_required
def view_transcript():
    if not current_user.is_student:
        return redirect(url_for('index'))
    return _render_transcript(current_user)


# ── Lecturer ──────────────────────────────────────────────────────────────────

@app.route('/lecturer')
@login_required
@lecturer_required
def lecturer_dashboard():
    if current_user.is_admin:
        assignments = CourseAssignment.query.filter_by(semester='1', year=2024).all()
    else:
        assignments = CourseAssignment.query.filter_by(
            lecturer_id=current_user.id, semester='1', year=2024).all()
    return render_template('lecturer/dashboard.html', assignments=assignments)


@app.route('/lecturer/grades/<int:course_id>', methods=['GET', 'POST'])
@login_required
@lecturer_required
def lecturer_grades(course_id):
    if not current_user.is_admin:
        ok = CourseAssignment.query.filter_by(
            lecturer_id=current_user.id, course_id=course_id,
            semester='1', year=2024).first()
        if not ok:
            abort(403)
    course    = Course.query.get_or_404(course_id)
    formula   = GradeFormula.query.filter_by(course_id=course_id).first()
    cw_weight = formula.classwork_weight if formula else 40
    fw_weight = formula.final_weight     if formula else 60

    if request.method == 'POST':
        for e in Enrollment.query.filter_by(course_id=course_id, semester='1', year=2024).all():
            old_cw    = e.classwork_grade
            old_final = e.final_grade
            cw_str    = request.form.get(f'cw_{e.id}', '').strip()
            final_str = request.form.get(f'final_{e.id}', '').strip()

            def _parse(s):
                if not s: return None
                v = float(s)
                if not 0 <= v <= 100: raise ValueError
                return v

            try:
                e.classwork_grade = _parse(cw_str)
                e.final_grade     = _parse(final_str)
                new_total = compute_grade(e)
                if new_total != e.grade:
                    action = '⚠ تعديل درجة بعد القفل' if e.grade_locked else 'تعديل درجة'
                    log_action(action, 'enrollment', e.id,
                               f'أعمال:{old_cw} امتحان:{old_final}',
                               f'أعمال:{e.classwork_grade} امتحان:{e.final_grade}')
                    e.grade  = new_total
                    if new_total is not None:
                        e.status = 'completed' if new_total >= 50 else 'enrolled'
            except ValueError:
                flash(f'قيمة غير صحيحة: {e.student.full_name}', 'error')

        db.session.commit()
        flash(f'تم حفظ درجات {course.name_ar}', 'success')
        return redirect(url_for('lecturer_grades', course_id=course_id))

    enrollments = Enrollment.query.filter_by(
        course_id=course_id, semester='1', year=2024).all()
    return render_template('lecturer/grades.html',
                           course=course, enrollments=enrollments,
                           cw_weight=cw_weight, fw_weight=fw_weight)


# ── Admin — Dashboard ─────────────────────────────────────────────────────────

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    return render_template('admin/dashboard.html',
        total_students=User.query.filter_by(role='student').count(),
        pending=User.query.filter_by(status='pending', role='student').count(),
        approved=User.query.filter_by(status='approved', role='student').count(),
        unpaid=User.query.filter_by(financial_status='unpaid', role='student').count(),
    )


# ── Admin — Students CRUD ─────────────────────────────────────────────────────

@app.route('/admin/students')
@login_required
@admin_required
def admin_students():
    students = User.query.filter_by(role='student').order_by(User.created_at.desc()).all()
    return render_template('admin/students.html', students=students)


@app.route('/admin/students/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_student(user_id):
    student = User.query.get_or_404(user_id)
    departments = Department.query.all()
    enrollments = Enrollment.query.filter_by(user_id=user_id).all()
    formulas    = {f.course_id: f for f in GradeFormula.query.all()}
    docs_by_type = {d.document_type: d
                    for d in Document.query.filter_by(user_id=user_id).all()}

    if request.method == 'POST':
        student.full_name = request.form.get('full_name', student.full_name).strip() or student.full_name
        student.status           = request.form.get('status', student.status)
        student.financial_status = request.form.get('financial_status', student.financial_status)
        dept_id = request.form.get('department_id', type=int)
        student.department_id = dept_id or None

        bd_str = request.form.get('birth_date', '')
        if bd_str:
            try:
                student.birth_date = datetime.strptime(bd_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        _save_docs_for('user_id', user_id)

        def _pg(s):
            if not s: return None
            v = float(s)
            if not 0 <= v <= 100: raise ValueError
            return v

        for e in enrollments:
            old_grade = e.grade
            try:
                cw_s    = request.form.get(f'cw_{e.id}', '').strip()
                final_s = request.form.get(f'final_{e.id}', '').strip()
                e.classwork_grade = _pg(cw_s)
                e.final_grade     = _pg(final_s)
                new_total = compute_grade(e)
                if new_total != old_grade:
                    log_action('تعديل درجة (مدير)', 'enrollment', e.id, old_grade, new_total)
                    e.grade  = new_total
                    if new_total is not None:
                        e.status = 'completed' if new_total >= 50 else 'enrolled'
            except ValueError:
                pass

        log_action('تعديل بيانات الطالب', 'user', user_id)
        db.session.commit()
        flash(f'تم تحديث بيانات {student.full_name}', 'success')
        return redirect(url_for('admin_students'))

    return render_template('admin/edit_student.html',
        student=student, departments=departments,
        enrollments=enrollments, docs_by_type=docs_by_type, formulas=formulas)


@app.route('/admin/students/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_student(user_id):
    student = User.query.get_or_404(user_id)
    if student.is_admin:
        flash('لا يمكن حذف مدير النظام', 'error')
        return redirect(url_for('admin_students'))
    name = student.full_name
    sid  = student.student_id

    # Cascade: nullify audit refs, wipe all relational data
    AuditLog.query.filter_by(user_id=user_id).update({'user_id': None})
    Enrollment.query.filter_by(user_id=user_id).delete()
    Document.query.filter_by(user_id=user_id).delete()
    CourseAssignment.query.filter_by(lecturer_id=user_id).delete()
    ClassSchedule.query.filter_by(lecturer_id=user_id).update({'lecturer_id': None})

    # Revoke re-registration ability: delete the pre-reg record so the admin
    # must explicitly re-add them before they can create a new account.
    pre_reg = PreRegisteredStudent.query.filter_by(student_id=sid).first()
    if pre_reg:
        Document.query.filter_by(pre_reg_id=pre_reg.id).delete()
        db.session.delete(pre_reg)

    log_action('حذف طالب (نهائي)', 'user', user_id, old_val=f'{sid}:{name}')
    db.session.delete(student)
    db.session.commit()
    flash(f'تم حذف {name} نهائياً وإلغاء تصريح إعادة التسجيل.', 'success')
    return redirect(url_for('admin_students'))


@app.route('/admin/verify/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def verify_student(user_id):
    user = User.query.get_or_404(user_id)
    user.status = 'approved'
    log_action('قبول الطالب', 'user', user_id, 'pending', 'approved')
    db.session.commit()
    flash(f'تمت الموافقة على {user.full_name}', 'success')
    return redirect(url_for('admin_students'))


@app.route('/admin/reject/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def reject_student(user_id):
    user = User.query.get_or_404(user_id)
    user.status = 'rejected'
    log_action('رفض الطالب', 'user', user_id, 'pending', 'rejected')
    db.session.commit()
    flash(f'تم رفض {user.full_name}', 'error')
    return redirect(url_for('admin_students'))


@app.route('/admin/financial/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def update_financial(user_id):
    user   = User.query.get_or_404(user_id)
    status = request.form.get('status')
    if status in ('paid', 'unpaid'):
        old = user.financial_status
        user.financial_status = status
        log_action('تحديث الحالة المالية', 'user', user_id, old, status)
        db.session.commit()
        flash(f'تم تحديث الحالة المالية لـ {user.full_name}', 'success')
    return redirect(url_for('admin_students'))


@app.route('/admin/role/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def change_role(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        flash('لا يمكن تغيير دور مدير النظام', 'error')
        return redirect(url_for('admin_students'))
    new_role = request.form.get('role')
    if new_role in ('student', 'lecturer'):
        old = user.role
        user.role = new_role
        log_action('تغيير الدور', 'user', user_id, old, new_role)
        db.session.commit()
        flash(f'تم تغيير دور {user.full_name}', 'success')
    return redirect(url_for('admin_students'))


# ── Admin — Lecturers CRUD ────────────────────────────────────────────────────

@app.route('/admin/lecturers')
@login_required
@admin_required
def admin_lecturers():
    lecturers = User.query.filter_by(role='lecturer').order_by(User.created_at.desc()).all()
    return render_template('admin/lecturers.html', lecturers=lecturers)


@app.route('/admin/lecturers/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_lecturer():
    departments = Department.query.all()
    if request.method == 'POST':
        sid       = request.form.get('student_id', '').strip()
        full_name = request.form.get('full_name', '').strip()
        email     = request.form.get('email', '').strip().lower()
        password  = request.form.get('password', '')
        dept_id   = request.form.get('department_id', type=int)
        bd_str    = request.form.get('birth_date', '')

        errors = []
        if not all([sid, full_name, email, password]):
            errors.append('جميع الحقول المطلوبة يجب ملؤها')
        if User.query.filter_by(email=email).first():
            errors.append('البريد الإلكتروني مستخدم مسبقاً')
        if User.query.filter_by(student_id=sid).first():
            errors.append('الرقم الوظيفي مستخدم مسبقاً')

        if errors:
            for e in errors:
                flash(e, 'error')
        else:
            birth_date = None
            if bd_str:
                try:
                    birth_date = datetime.strptime(bd_str, '%Y-%m-%d').date()
                except ValueError:
                    pass

            lec = User(
                student_id=sid, full_name=full_name, email=email,
                password_hash=generate_password_hash(password),
                role='lecturer', status='approved', financial_status='paid',
                department_id=dept_id or None, birth_date=birth_date,
            )
            db.session.add(lec)
            db.session.flush()

            prof_path = save_upload(request.files.get('profile_image'), 'profiles', ALLOWED_IMAGES)
            if prof_path:
                lec.profile_image = prof_path
            _save_docs_for('user_id', lec.id)

            log_action('إضافة محاضر', 'user', lec.id, new_val=email)
            db.session.commit()
            flash(f'تمت إضافة المحاضر {full_name}', 'success')
            return redirect(url_for('admin_lecturers'))

    return render_template('admin/lecturer_form.html',
        lecturer=None, departments=departments, title='إضافة محاضر جديد', docs_by_type={})


@app.route('/admin/lecturers/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_lecturer(user_id):
    lecturer    = User.query.filter_by(id=user_id, role='lecturer').first_or_404()
    departments = Department.query.all()
    docs_by_type = {d.document_type: d
                    for d in Document.query.filter_by(user_id=user_id).all()}

    if request.method == 'POST':
        lecturer.full_name     = request.form.get('full_name', lecturer.full_name).strip() or lecturer.full_name
        lecturer.department_id = request.form.get('department_id', type=int) or None
        new_pw = request.form.get('password', '').strip()
        if new_pw:
            lecturer.password_hash = generate_password_hash(new_pw)

        bd_str = request.form.get('birth_date', '')
        if bd_str:
            try:
                lecturer.birth_date = datetime.strptime(bd_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        prof_path = save_upload(request.files.get('profile_image'), 'profiles', ALLOWED_IMAGES)
        if prof_path:
            lecturer.profile_image = prof_path
        _save_docs_for('user_id', user_id)

        log_action('تعديل بيانات محاضر', 'user', user_id)
        db.session.commit()
        flash(f'تم تحديث بيانات {lecturer.full_name}', 'success')
        return redirect(url_for('admin_lecturers'))

    return render_template('admin/lecturer_form.html',
        lecturer=lecturer, departments=departments,
        docs_by_type=docs_by_type, title=f'تعديل: {lecturer.full_name}')


@app.route('/admin/lecturers/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_lecturer(user_id):
    lecturer = User.query.filter_by(id=user_id, role='lecturer').first_or_404()
    name = lecturer.full_name

    # Cascade: nullify audit refs, wipe all relational data
    AuditLog.query.filter_by(user_id=user_id).update({'user_id': None})
    Document.query.filter_by(user_id=user_id).delete()
    CourseAssignment.query.filter_by(lecturer_id=user_id).delete()
    # NULL the lecturer_id on schedule slots — keeps the timetable intact but
    # removes the FK reference that would otherwise block deletion.
    ClassSchedule.query.filter_by(lecturer_id=user_id).update({'lecturer_id': None})

    log_action('حذف محاضر (نهائي)', 'user', user_id, old_val=name)
    db.session.delete(lecturer)
    db.session.commit()
    flash(f'تم حذف المحاضر {name} نهائياً وتحرير جميع ارتباطاته.', 'success')
    return redirect(url_for('admin_lecturers'))


# ── Admin — Departments CRUD ──────────────────────────────────────────────────

@app.route('/admin/departments', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_departments():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            name_ar = request.form.get('name_ar', '').strip()
            name_en = request.form.get('name_en', '').strip()
            if not name_ar or not name_en:
                flash('جميع الحقول مطلوبة', 'error')
            else:
                db.session.add(Department(name_ar=name_ar, name_en=name_en))
                db.session.commit()
                flash(f'تمت إضافة قسم {name_ar}', 'success')

        elif action == 'edit':
            dept = Department.query.get_or_404(request.form.get('dept_id', type=int))
            dept.name_ar = request.form.get('name_ar', dept.name_ar).strip() or dept.name_ar
            dept.name_en = request.form.get('name_en', dept.name_en).strip() or dept.name_en
            db.session.commit()
            flash('تم تحديث القسم', 'success')

        elif action == 'delete':
            dept   = Department.query.get_or_404(request.form.get('dept_id', type=int))
            c_cnt  = dept.courses.count()
            u_cnt  = User.query.filter_by(department_id=dept.id).count()
            pr_cnt = PreRegisteredStudent.query.filter_by(department_id=dept.id).count()
            if c_cnt + u_cnt + pr_cnt > 0:
                flash(f'لا يمكن الحذف: القسم مرتبط بـ {c_cnt} مادة و {u_cnt + pr_cnt} شخص', 'error')
            else:
                name = dept.name_ar
                db.session.delete(dept)
                db.session.commit()
                flash(f'تم حذف قسم {name}', 'success')

    departments = Department.query.all()
    return render_template('admin/departments.html', departments=departments)


# ── Admin — Batch Grades ──────────────────────────────────────────────────────

@app.route('/admin/batch-grades', methods=['GET', 'POST'])
@login_required
@admin_required
def batch_grades():
    if request.method == 'POST':
        try:
            adjustment = float(request.form.get('adjustment', 0))
        except ValueError:
            flash('قيمة التعديل غير صحيحة', 'error')
            return redirect(url_for('batch_grades'))

        scope     = request.form.get('scope', 'all')
        course_id = request.form.get('course_id', type=int)

        query = Enrollment.query.filter(Enrollment.grade.isnot(None))
        if scope == 'course' and course_id:
            query = query.filter_by(course_id=course_id)

        count = 0
        for e in query.all():
            old_grade = e.grade
            new_grade = round(min(100.0, max(0.0, e.grade + adjustment)), 2)
            e.grade   = new_grade
            e.status  = 'completed' if new_grade >= 50 else 'enrolled'
            log_action(f'تعديل جماعي ({adjustment:+.1f})', 'enrollment',
                       e.id, old_grade, new_grade)
            count += 1

        db.session.commit()
        sign = '+' if adjustment >= 0 else ''
        flash(f'تم تعديل {count} درجة ({sign}{adjustment})', 'success')
        return redirect(url_for('batch_grades'))

    courses = Course.query.all()
    return render_template('admin/batch_grades.html', courses=courses)


# ── Admin — Pre-register ──────────────────────────────────────────────────────

@app.route('/admin/pre-register', methods=['GET', 'POST'])
@login_required
@admin_required
def pre_register():
    departments = Department.query.all()
    if request.method == 'POST':
        sid     = request.form.get('student_id', '').strip()
        name    = request.form.get('full_name', '').strip()
        dept_id = request.form.get('department_id', type=int)
        bd_str  = request.form.get('birth_date', '')

        if not sid or not name:
            flash('رقم القيد والاسم مطلوبان', 'error')
        elif PreRegisteredStudent.query.filter_by(student_id=sid).first():
            flash('رقم القيد موجود مسبقاً', 'error')
        else:
            birth_date = None
            if bd_str:
                try:
                    birth_date = datetime.strptime(bd_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            student = PreRegisteredStudent(
                student_id=sid, full_name=name,
                department_id=dept_id or None, birth_date=birth_date)
            db.session.add(student)
            db.session.flush()
            _save_docs_for('pre_reg_id', student.id)
            log_action('إضافة طالب مسجل مسبقاً', 'pre_registered', new_val=f'{sid}:{name}')
            db.session.commit()
            flash(f'تمت إضافة {name}', 'success')

    students = PreRegisteredStudent.query.order_by(
        PreRegisteredStudent.created_at.desc()).all()
    return render_template('admin/pre_register.html', students=students, departments=departments)


@app.route('/admin/preregistered/<int:pre_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_pre_registered(pre_id):
    student = PreRegisteredStudent.query.get_or_404(pre_id)
    name = student.full_name
    sid  = student.student_id

    # Delete documents linked to this pre-registration record only
    Document.query.filter_by(pre_reg_id=pre_id).delete()

    log_action('حذف من التسجيل المسبق', 'pre_registered', pre_id,
               old_val=f'{sid}:{name}')
    db.session.delete(student)
    db.session.commit()

    # Inform admin if an active user account already exists for this student_id
    existing_user = User.query.filter_by(student_id=sid).first()
    if existing_user:
        flash(
            f'تم حذف "{name}" من قائمة التسجيل المسبق. '
            f'تنبيه: لديه حساب نشط في النظام لم يتأثر بالحذف.',
            'warning'
        )
    else:
        flash(f'تم حذف "{name}" من قائمة التسجيل المسبق بنجاح.', 'success')
    return redirect(url_for('pre_register'))


# ── Admin — Import / Export / Analytics / Audit ───────────────────────────────

@app.route('/admin/import', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_import():
    results = None
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('يرجى اختيار ملف', 'error')
            return render_template('admin/import.html', results=None)

        fname = file.filename.lower()
        imported = skipped = 0
        errors   = []

        try:
            if fname.endswith(('.xlsx', '.xls')):
                wb   = openpyxl.load_workbook(file, data_only=True)
                ws   = wb.active
                rows = [(ws.cell(r, 1).value, ws.cell(r, 2).value)
                        for r in range(2, ws.max_row + 1)]
            elif fname.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader  = csv.reader(io.StringIO(content))
                next(reader, None)
                rows = list(reader)
            else:
                flash('يُدعم فقط xlsx و csv', 'error')
                return render_template('admin/import.html', results=None)

            for i, row in enumerate(rows, 2):
                sid  = str(row[0]).strip() if row and row[0] else ''
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not sid or not name:
                    errors.append(f'الصف {i}: بيانات ناقصة')
                    continue
                if PreRegisteredStudent.query.filter_by(student_id=sid).first():
                    skipped += 1
                    continue
                db.session.add(PreRegisteredStudent(student_id=sid, full_name=name))
                imported += 1

            db.session.commit()
            log_action('استيراد قائمة طلاب', 'pre_registered', new_val=f'added={imported}')
            db.session.commit()
            results = {'imported': imported, 'skipped': skipped, 'errors': errors}
            flash(f'تم الاستيراد: {imported} جديد، تخطي: {skipped} مكرر', 'success')

        except Exception as exc:
            flash(f'خطأ في الملف: {exc}', 'error')

    return render_template('admin/import.html', results=results)


@app.route('/admin/import/template')
@login_required
@admin_required
def import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الطلاب'
    ws.append(['student_id', 'full_name'])
    ws.append(['218006', 'مثال اسم الطالب'])
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='import_template.xlsx')


@app.route('/admin/analytics')
@login_required
@admin_required
def admin_analytics():
    total_students  = User.query.filter_by(role='student').count()
    all_enrollments = Enrollment.query.filter_by(semester='1', year=2024).all()
    graded   = [e for e in all_enrollments if e.grade is not None]
    passed   = sum(1 for e in graded if e.grade >= 50)
    failed   = len(graded) - passed
    pass_rate = round(passed / len(graded) * 100, 1) if graded else 0

    at_risk_map = {}
    for e in all_enrollments:
        is_at_risk = (
            (e.classwork_grade is not None and e.classwork_grade < 40) or
            (e.grade is not None and e.grade < 50))
        if is_at_risk:
            if e.user_id not in at_risk_map:
                at_risk_map[e.user_id] = {
                    'student': e.student, 'courses': [],
                    'warning': 'high' if (e.grade and e.grade < 40) else 'medium'}
            at_risk_map[e.user_id]['courses'].append(e.course.name_ar)
    at_risk = list(at_risk_map.values())

    depts       = Department.query.all()
    dept_labels = [d.name_ar for d in depts]
    dept_counts = [User.query.filter_by(role='student', department_id=d.id).count() for d in depts]

    monthly = db.session.query(
        extract('month', User.created_at).label('m'),
        func.count(User.id).label('c')
    ).filter(User.role == 'student').group_by('m').order_by('m').all()

    mnames  = {1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',5:'مايو',6:'يونيو',
               7:'يوليو',8:'أغسطس',9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'}
    m_labels = [mnames.get(int(r.m), '') for r in monthly]
    m_counts = [r.c for r in monthly]

    return render_template('admin/analytics.html',
        total_students=total_students, enrolled_count=len(all_enrollments),
        pass_rate=pass_rate, at_risk_count=len(at_risk), at_risk=at_risk,
        passed=passed, failed=failed,
        dept_labels=dept_labels, dept_counts=dept_counts,
        m_labels=m_labels, m_counts=m_counts)


@app.route('/admin/audit-log')
@login_required
@admin_required
def admin_audit_log():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(300).all()
    return render_template('admin/audit_log.html', logs=logs)


@app.route('/admin/export/excel')
@login_required
@admin_required
def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الطلاب'
    hdr_fill = PatternFill('solid', fgColor='1e3a5f')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    headers  = ['رقم القيد', 'الاسم', 'البريد', 'الحالة', 'الحالة المالية', 'القسم', 'تاريخ الميلاد', 'تاريخ التسجيل']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    status_map   = {'approved': 'مقبول', 'pending': 'قيد المراجعة', 'rejected': 'مرفوض'}
    financial_map = {'paid': 'مدفوع', 'unpaid': 'غير مدفوع'}
    for row, s in enumerate(User.query.filter_by(role='student').all(), 2):
        ws.cell(row, 1, s.student_id)
        ws.cell(row, 2, s.full_name)
        ws.cell(row, 3, s.email)
        ws.cell(row, 4, status_map.get(s.status, s.status))
        ws.cell(row, 5, financial_map.get(s.financial_status, s.financial_status))
        ws.cell(row, 6, s.department.name_ar if s.department else '')
        ws.cell(row, 7, s.birth_date.strftime('%Y-%m-%d') if s.birth_date else '')
        ws.cell(row, 8, s.created_at.strftime('%Y-%m-%d'))
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = (
            max(len(str(c.value or '')) for c in col_cells) + 4)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    log_action('تصدير Excel', 'export'); db.session.commit()
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'students_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/admin/transcript/<int:user_id>')
@login_required
@admin_required
def admin_transcript(user_id):
    return _render_transcript(User.query.get_or_404(user_id))


@app.route('/admin/assign-lecturer', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_lecturer():
    if request.method == 'POST':
        lecturer_id = request.form.get('lecturer_id', type=int)
        course_id   = request.form.get('course_id', type=int)
        if lecturer_id and course_id:
            exists = CourseAssignment.query.filter_by(
                lecturer_id=lecturer_id, course_id=course_id,
                semester='1', year=2024).first()
            if not exists:
                db.session.add(CourseAssignment(
                    lecturer_id=lecturer_id, course_id=course_id,
                    semester='1', year=2024))
                log_action('تعيين محاضر', 'course_assignment',
                           course_id, new_val=f'lecturer={lecturer_id}')
                db.session.commit()
                flash('تم التعيين بنجاح', 'success')
            else:
                flash('المحاضر معين لهذه المادة مسبقاً', 'warning')
        return redirect(url_for('assign_lecturer'))

    lecturers   = User.query.filter_by(role='lecturer').all()
    courses     = Course.query.all()
    assignments = CourseAssignment.query.filter_by(semester='1', year=2024).all()
    return render_template('admin/assign_lecturer.html',
        lecturers=lecturers, courses=courses, assignments=assignments)


# ── Admin — Staff Management (Full-Access only) ───────────────────────────────

@app.route('/admin/staff', methods=['GET', 'POST'])
@login_required
@full_access_required
def admin_staff():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            full_name   = request.form.get('full_name', '').strip()
            employee_id = request.form.get('employee_id', '').strip()
            email       = request.form.get('email', '').strip().lower()
            password    = request.form.get('password', '')
            permissions = request.form.get('permissions', 'full_access')

            errors = []
            if not all([full_name, employee_id, email, password]):
                errors.append('جميع الحقول المطلوبة يجب ملؤها')
            if User.query.filter_by(email=email).first():
                errors.append('البريد الإلكتروني مستخدم مسبقاً')
            if User.query.filter_by(student_id=employee_id).first():
                errors.append('الرقم الوظيفي مستخدم مسبقاً')
            if len(password) < 8:
                errors.append('كلمة المرور يجب أن تكون 8 أحرف على الأقل')

            if errors:
                for e in errors:
                    flash(e, 'error')
            else:
                new_admin = User(
                    student_id=employee_id, full_name=full_name, email=email,
                    password_hash=generate_password_hash(password),
                    role='admin', status='approved', financial_status='paid',
                    admin_permissions=permissions,
                )
                db.session.add(new_admin)
                db.session.flush()
                log_action('إضافة مدير', 'user', new_admin.id,
                           new_val=f'{email}:{permissions}')
                db.session.commit()
                flash(f'تمت إضافة المدير {full_name}', 'success')
                return redirect(url_for('admin_staff'))

        elif action == 'edit':
            admin_id   = request.form.get('admin_id', type=int)
            admin_user = User.query.filter_by(id=admin_id, role='admin').first_or_404()
            if admin_user.id == current_user.id:
                flash('لا يمكن تعديل حسابك الخاص من هنا', 'error')
                return redirect(url_for('admin_staff'))

            admin_user.full_name = (
                request.form.get('full_name', admin_user.full_name).strip()
                or admin_user.full_name
            )
            old_perms = admin_user.admin_permissions
            new_perms = request.form.get('permissions', old_perms)
            admin_user.admin_permissions = new_perms

            new_pw = request.form.get('password', '').strip()
            if new_pw:
                if len(new_pw) < 8:
                    flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
                    return redirect(url_for('admin_staff'))
                admin_user.password_hash = generate_password_hash(new_pw)

            log_action('تعديل مدير', 'user', admin_id, old_perms, new_perms)
            db.session.commit()
            flash(f'تم تحديث بيانات {admin_user.full_name}', 'success')
            return redirect(url_for('admin_staff'))

    admins = User.query.filter_by(role='admin').order_by(User.created_at).all()
    return render_template('admin/staff.html', admins=admins)


@app.route('/admin/staff/<int:admin_id>/delete', methods=['POST'])
@login_required
@full_access_required
def delete_admin_staff(admin_id):
    if admin_id == current_user.id:
        flash('لا يمكنك حذف حسابك الخاص', 'error')
        return redirect(url_for('admin_staff'))
    admin_user = User.query.filter_by(id=admin_id, role='admin').first_or_404()
    name = admin_user.full_name
    AuditLog.query.filter_by(user_id=admin_id).update({'user_id': None})
    log_action('حذف مدير', 'user', admin_id, old_val=name)
    db.session.delete(admin_user)
    db.session.commit()
    flash(f'تم حذف المدير {name}', 'success')
    return redirect(url_for('admin_staff'))


# ── Admin — Global Enrollment Control ────────────────────────────────────────

@app.route('/admin/enrollment-control')
@login_required
@admin_required
def admin_enrollment_control():
    q = request.args.get('q', '').strip()
    students = []
    student_enrollments = {}

    if q:
        students = User.query.filter(
            User.role == 'student',
            db.or_(
                User.full_name.ilike(f'%{q}%'),
                User.student_id.ilike(f'%{q}%'),
            )
        ).order_by(User.full_name).all()
        for s in students:
            student_enrollments[s.id] = Enrollment.query.filter_by(
                user_id=s.id, semester='1', year=2024, dropped_by_admin=False
            ).all()

    return render_template('admin/enrollment_control.html',
                           students=students,
                           student_enrollments=student_enrollments, q=q)


@app.route('/admin/enrollment-control/drop/<int:enrollment_id>', methods=['POST'])
@login_required
@admin_required
def admin_drop_enrollment(enrollment_id):
    enrollment = Enrollment.query.get_or_404(enrollment_id)
    student    = User.query.get(enrollment.user_id)
    course     = Course.query.get(enrollment.course_id)
    old_status = enrollment.status
    enrollment.dropped_by_admin = True
    enrollment.status = 'admin_dropped'
    log_action('إسقاط مادة (إداري)', 'enrollment', enrollment_id,
               old_val=f'{student.full_name}:{course.code}:{old_status}',
               new_val='admin_dropped')
    db.session.commit()
    flash(f'تم إسقاط {course.name_ar} للطالب {student.full_name}', 'success')
    q = request.form.get('q', '')
    return redirect(url_for('admin_enrollment_control', q=q))


# ── Admin — Branding / System Settings ───────────────────────────────────────

MAX_LOGO_BYTES = 5 * 1024 * 1024  # 5 MB


@app.route('/admin/branding', methods=['GET', 'POST'])
@login_required
@full_access_required
def admin_branding():
    settings = SystemSettings.query.first()

    if request.method == 'POST':
        logo_file = request.files.get('logo')
        if not logo_file or not logo_file.filename:
            flash('يرجى اختيار ملف صورة', 'error')
            return render_template('admin/branding.html', settings=settings)

        ext = logo_file.filename.rsplit('.', 1)[-1].lower() if '.' in logo_file.filename else ''
        if ext not in ALLOWED_IMAGES:
            flash('صيغة الملف غير مدعومة. يُقبل فقط: jpg, jpeg, png, gif, webp', 'error')
            return render_template('admin/branding.html', settings=settings)

        # Size guard (read into memory for check, then re-seek)
        logo_file.seek(0, 2)
        file_size = logo_file.tell()
        logo_file.seek(0)
        if file_size > MAX_LOGO_BYTES:
            flash('حجم الملف يتجاوز 5 ميجابايت', 'error')
            return render_template('admin/branding.html', settings=settings)

        # Remove old custom logo file if it exists
        if settings and settings.logo_path and settings.logo_path.startswith('uploads/branding/'):
            old_abs = os.path.join(app.static_folder, settings.logo_path)
            if os.path.exists(old_abs):
                try:
                    os.remove(old_abs)
                except OSError:
                    pass

        # Save new file
        save_dir = os.path.join(app.static_folder, 'uploads', 'branding')
        os.makedirs(save_dir, exist_ok=True)
        unique_name = f"logo_{uuid.uuid4().hex}.{ext}"
        logo_file.save(os.path.join(save_dir, unique_name))
        new_path = f'uploads/branding/{unique_name}'

        if settings:
            settings.logo_path = new_path
        else:
            settings = SystemSettings(logo_path=new_path)
            db.session.add(settings)

        log_action('تحديث شعار النظام [superadmin]', 'system_settings',
                   entity_id=1, new_val=new_path)
        db.session.commit()
        flash('تم تحديث شعار النظام بنجاح وسيظهر فوراً في جميع الصفحات', 'success')
        return redirect(url_for('admin_branding'))

    return render_template('admin/branding.html', settings=settings)


# ── Admin — Course Management ─────────────────────────────────────────────────

@app.route('/admin/courses', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_courses():
    departments = Department.query.order_by(Department.name_ar).all()
    all_courses = Course.query.order_by(Course.code).all()

    if request.method == 'POST':
        action    = request.form.get('action')
        course_id = request.form.get('course_id', type=int)
        code      = request.form.get('code', '').strip().upper()
        name_ar   = request.form.get('name_ar', '').strip()
        name_en   = request.form.get('name_en', '').strip()
        credits   = request.form.get('credits', 3, type=int)
        dept_id   = request.form.get('department_id', type=int)
        prereq_id = request.form.get('prerequisite_id', type=int) or None

        if action == 'add':
            errors = []
            if not all([code, name_ar, name_en, dept_id]):
                errors.append('الرمز والاسمان والقسم حقول مطلوبة')
            elif Course.query.filter_by(code=code).first():
                errors.append('رمز المادة مستخدم مسبقاً')
            if errors:
                for e in errors: flash(e, 'error')
            else:
                c = Course(code=code, name_ar=name_ar, name_en=name_en,
                           credits=credits, department_id=dept_id,
                           prerequisite_id=prereq_id)
                db.session.add(c)
                log_action('إضافة مادة', 'course', new_val=f'{code}:{name_ar}')
                db.session.commit()
                flash(f'تمت إضافة المادة {name_ar}', 'success')
                return redirect(url_for('admin_courses'))

        elif action == 'edit' and course_id:
            c   = Course.query.get_or_404(course_id)
            dup = Course.query.filter_by(code=code).first()
            if dup and dup.id != course_id:
                flash('رمز المادة مستخدم من مادة أخرى', 'error')
            else:
                old = c.code
                c.code = code; c.name_ar = name_ar; c.name_en = name_en
                c.credits = credits; c.department_id = dept_id
                c.prerequisite_id = prereq_id if prereq_id != course_id else None
                log_action('تعديل مادة', 'course', course_id, old, code)
                db.session.commit()
                flash(f'تم تحديث المادة {name_ar}', 'success')
                return redirect(url_for('admin_courses'))

    dept_filter = request.args.get('dept', type=int)
    q = request.args.get('q', '').strip()
    filtered = all_courses
    if dept_filter:
        filtered = [c for c in filtered if c.department_id == dept_filter]
    if q:
        ql = q.lower()
        filtered = [c for c in filtered
                    if ql in c.code.lower() or ql in c.name_ar or ql in c.name_en.lower()]
    return render_template('admin/courses.html',
                           courses=filtered, all_courses=all_courses,
                           departments=departments, dept_filter=dept_filter, q=q)


@app.route('/admin/courses/<int:course_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_course(course_id):
    c = Course.query.get_or_404(course_id)
    if c.enrollments.count() > 0:
        flash(f'لا يمكن حذف {c.name_ar}: يوجد {c.enrollments.count()} طالب مسجل', 'error')
        return redirect(url_for('admin_courses'))
    dependents = Course.query.filter_by(prerequisite_id=course_id).all()
    if dependents:
        flash(f'لا يمكن حذف {c.name_ar}: مطلب سابق لـ '
              f'{", ".join(x.name_ar for x in dependents)}', 'error')
        return redirect(url_for('admin_courses'))
    name = c.name_ar
    ClassSchedule.query.filter_by(course_id=course_id).delete()
    CourseAssignment.query.filter_by(course_id=course_id).delete()
    log_action('حذف مادة', 'course', course_id, old_val=f'{c.code}:{name}')
    db.session.delete(c)
    db.session.commit()
    flash(f'تم حذف المادة {name}', 'success')
    return redirect(url_for('admin_courses'))


# ── Admin — Classroom Management ─────────────────────────────────────────────

@app.route('/admin/classrooms', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_classrooms():
    if request.method == 'POST':
        action    = request.form.get('action')
        room_id   = request.form.get('room_id', type=int)
        name      = request.form.get('name', '').strip()
        capacity  = request.form.get('capacity', 30, type=int)
        room_type = request.form.get('room_type', 'lecture')

        if action == 'add':
            if not name:
                flash('اسم القاعة مطلوب', 'error')
            elif Classroom.query.filter_by(name=name).first():
                flash('اسم القاعة مستخدم مسبقاً', 'error')
            else:
                r = Classroom(name=name, capacity=capacity, room_type=room_type)
                db.session.add(r)
                log_action('إضافة قاعة', 'classroom', new_val=name)
                db.session.commit()
                flash(f'تمت إضافة القاعة {name}', 'success')
                return redirect(url_for('admin_classrooms'))

        elif action == 'edit' and room_id:
            r   = Classroom.query.get_or_404(room_id)
            dup = Classroom.query.filter_by(name=name).first()
            if dup and dup.id != room_id:
                flash('اسم القاعة مستخدم من قاعة أخرى', 'error')
            else:
                old = r.name
                r.name = name; r.capacity = capacity; r.room_type = room_type
                log_action('تعديل قاعة', 'classroom', room_id, old, name)
                db.session.commit()
                flash(f'تم تحديث القاعة {name}', 'success')
                return redirect(url_for('admin_classrooms'))

    classrooms = Classroom.query.order_by(Classroom.name).all()
    return render_template('admin/classrooms.html', classrooms=classrooms)


@app.route('/admin/classrooms/<int:room_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_classroom(room_id):
    r = Classroom.query.get_or_404(room_id)
    if r.schedules.count() > 0:
        flash(f'لا يمكن حذف {r.name}: مستخدمة في الجداول', 'error')
        return redirect(url_for('admin_classrooms'))
    name = r.name
    log_action('حذف قاعة', 'classroom', room_id, old_val=name)
    db.session.delete(r)
    db.session.commit()
    flash(f'تم حذف القاعة {name}', 'success')
    return redirect(url_for('admin_classrooms'))


# ── Admin — Class Scheduling ──────────────────────────────────────────────────

DAYS_AR    = ['الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
TIME_SLOTS = ['08:00', '09:00', '10:00', '11:00', '12:00',
              '13:00', '14:00', '15:00', '16:00', '17:00']


def _check_schedule_conflicts(lecturer_id, classroom_id, day, start, end, exclude_id=None):
    msgs = []
    base = ClassSchedule.query.filter(
        ClassSchedule.day_of_week == day,
        ClassSchedule.start_time < end,
        ClassSchedule.end_time > start,
    )
    if exclude_id:
        base = base.filter(ClassSchedule.id != exclude_id)
    if lecturer_id:
        for s in base.filter(ClassSchedule.lecturer_id == lecturer_id).all():
            msgs.append(f'تعارض محاضر مع مادة "{s.course.name_ar}" ({s.start_time}–{s.end_time})')
    if classroom_id:
        for s in base.filter(ClassSchedule.classroom_id == classroom_id).all():
            msgs.append(f'تعارض قاعة مع مادة "{s.course.name_ar}" ({s.start_time}–{s.end_time})')
    return msgs


@app.route('/admin/schedule', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_schedule():
    sem     = request.args.get('semester', '1')
    yr      = request.args.get('year', 2024, type=int)
    dept_id = request.args.get('department_id', type=int)

    if request.method == 'POST':
        p_course_id    = request.form.get('course_id', type=int)
        p_lecturer_id  = request.form.get('lecturer_id', type=int) or None
        p_classroom_id = request.form.get('classroom_id', type=int) or None
        p_day          = request.form.get('day_of_week', '')
        p_start        = request.form.get('start_time', '')
        p_end          = request.form.get('end_time', '')
        p_sem          = request.form.get('semester', '1')
        p_yr           = request.form.get('year', 2024, type=int)

        errors = []
        if not all([p_course_id, p_day, p_start, p_end]):
            errors.append('المادة واليوم والوقت حقول مطلوبة')
        elif p_start >= p_end:
            errors.append('وقت البداية يجب أن يكون قبل وقت النهاية')
        else:
            errors.extend(_check_schedule_conflicts(
                p_lecturer_id, p_classroom_id, p_day, p_start, p_end))

        if errors:
            for e in errors: flash(e, 'error')
        else:
            db.session.add(ClassSchedule(
                course_id=p_course_id, lecturer_id=p_lecturer_id,
                classroom_id=p_classroom_id, day_of_week=p_day,
                start_time=p_start, end_time=p_end,
                semester=p_sem, year=p_yr,
            ))
            log_action('إضافة حصة', 'schedule',
                       new_val=f'{p_day} {p_start}-{p_end}')
            db.session.commit()
            flash('تمت إضافة الحصة إلى الجدول', 'success')
        return redirect(url_for('admin_schedule',
                                semester=p_sem, year=p_yr, department_id=dept_id))

    q = ClassSchedule.query.filter_by(semester=sem, year=yr)
    if dept_id:
        cids = [c.id for c in Course.query.filter_by(department_id=dept_id).all()]
        q = q.filter(ClassSchedule.course_id.in_(cids))
    schedules_raw = q.order_by(ClassSchedule.day_of_week, ClassSchedule.start_time).all()
    # Drop orphaned rows (course deleted after schedule was created)
    schedules = [s for s in schedules_raw
                 if s.course is not None and s.start_time and s.end_time]

    grid = {day: {slot: [] for slot in TIME_SLOTS} for day in DAYS_AR}
    for s in schedules:
        if s.day_of_week in grid:
            for slot in TIME_SLOTS:
                if s.start_time <= slot < s.end_time:
                    grid[s.day_of_week][slot].append(s)

    courses     = Course.query.order_by(Course.code).all()
    lecturers   = User.query.filter_by(role='lecturer').order_by(User.full_name).all()
    classrooms  = Classroom.query.order_by(Classroom.name).all()
    departments = Department.query.order_by(Department.name_ar).all()
    return render_template('admin/schedule.html',
        schedules=schedules, grid=grid, time_slots=TIME_SLOTS, days=DAYS_AR,
        courses=courses, lecturers=lecturers, classrooms=classrooms,
        departments=departments, sem=sem, yr=yr, dept_id=dept_id)


@app.route('/admin/schedule/<int:schedule_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_schedule(schedule_id):
    s = ClassSchedule.query.get_or_404(schedule_id)
    sem = s.semester; yr = s.year
    course_name = s.course.name_ar if s.course else '(مادة محذوفة)'
    log_action('حذف حصة', 'schedule', schedule_id,
               old_val=f'{course_name} {s.day_of_week} {s.start_time}')
    db.session.delete(s)
    db.session.commit()
    flash('تم حذف الحصة من الجدول', 'success')
    return redirect(url_for('admin_schedule', semester=sem, year=yr))


@app.route('/admin/schedule/export')
@login_required
@admin_required
def timetable_export():
    sem     = request.args.get('semester', '1')
    yr      = request.args.get('year', 2024, type=int)
    dept_id = request.args.get('department_id', type=int)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    depts = ([Department.query.get(dept_id)] if dept_id
             else Department.query.order_by(Department.name_ar).all())
    hdr_fill = PatternFill('solid', fgColor='1e3a5f')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)

    for dept in depts:
        if not dept:
            continue
        ws = wb.create_sheet(title=dept.name_ar[:31])
        ws.cell(1, 1, 'الوقت').fill = hdr_fill
        ws.cell(1, 1).font = hdr_font
        for col, day in enumerate(DAYS_AR, 2):
            cell = ws.cell(1, col, day)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')

        cids = [c.id for c in Course.query.filter_by(department_id=dept.id).all()]
        dept_scheds = ClassSchedule.query.filter(
            ClassSchedule.semester == sem,
            ClassSchedule.year == yr,
            ClassSchedule.course_id.in_(cids),
        ).all()

        grid = {day: {slot: [] for slot in TIME_SLOTS} for day in DAYS_AR}
        for s in dept_scheds:
            if s.day_of_week in grid:
                for slot in TIME_SLOTS:
                    if s.start_time <= slot < s.end_time:
                        grid[s.day_of_week][slot].append(s)

        for row, slot in enumerate(TIME_SLOTS, 2):
            ws.cell(row, 1, slot)
            for col, day in enumerate(DAYS_AR, 2):
                entries = grid[day][slot]
                if entries:
                    parts = []
                    for e in entries:
                        txt = e.course.name_ar
                        if e.classroom:
                            txt += f'\n({e.classroom.name})'
                        if e.lecturer:
                            txt += f'\n{e.lecturer.full_name}'
                        parts.append(txt)
                    cell = ws.cell(row, col, '\n---\n'.join(parts))
                    cell.alignment = Alignment(wrap_text=True,
                                               horizontal='center', vertical='top')

        ws.column_dimensions['A'].width = 8
        for col in range(2, len(DAYS_AR) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 28
        for row in range(1, len(TIME_SLOTS) + 2):
            ws.row_dimensions[row].height = 55

    if not wb.sheetnames:
        wb.create_sheet('الجدول')
    out = io.BytesIO(); wb.save(out); out.seek(0)
    log_action('تصدير جدول دراسي', 'schedule', new_val=f'{sem}/{yr}')
    db.session.commit()
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'timetable_{sem}_{yr}.xlsx')



# --- بداية الجزء الجاهز للنسخ واللصق (نسخة مصححة 100%) ---

from models import User, PreRegisteredStudent, Department, Course
from werkzeug.security import generate_password_hash
import os

with app.app_context():

    # DROP TABLE lines removed — db.create_all() below handles schema safely.

    db.create_all()

    ADMIN_EMAIL    = os.environ.get('ADMIN_EMAIL')
    ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD')
    MY_STUDENT_ID  = os.environ.get('MY_STUDENT_ID')
    MY_FULL_NAME   = os.environ.get('MY_FULL_NAME')

    # ═══════════════════════════════════════════════════════════════
    #  SEED — 7 Departments × 10 Courses  (idempotent upsert)
    # ═══════════════════════════════════════════════════════════════
    SEED = [
        {
            'name_ar': 'القانون', 'name_en': 'Law',
            'courses': [
                ('LAW101', 'المدخل لدراسة القانون',                  'Introduction to Law',                3),
                ('LAW102', 'القانون الدستوري',                        'Constitutional Law',                 3),
                ('LAW103', 'القانون الدولي العام',                    'Public International Law',           3),
                ('LAW104', 'مصادر الالتزام',                          'Sources of Obligation',              4),
                ('LAW105', 'أحكام الالتزام',                          'Rules of Obligation',                4),
                ('LAW106', 'القانون الجنائي العام',                   'General Criminal Law',               3),
                ('LAW107', 'القانون الجنائي الخاص',                   'Special Criminal Law',               3),
                ('LAW108', 'الشريعة الإسلامية: أحكام الأسرة',         'Islamic Sharia: Family Law',         3),
                ('LAW109', 'الشريعة الإسلامية: المواريث',             'Islamic Sharia: Inheritance Law',    3),
                ('LAW110', 'القانون الإداري',                         'Administrative Law',                 3),
            ],
        },
        {
            'name_ar': 'تصميم داخلي', 'name_en': 'Interior Design',
            'courses': [
                ('INT101', 'أساسيات التصميم الداخلي',                 'Interior Design Fundamentals',       3),
                ('INT102', 'الرسم الهندسي والمعماري',                  'Technical and Architectural Drawing',4),
                ('INT103', 'نظرية الألوان وتطبيقاتها',                 'Color Theory and Applications',      3),
                ('INT104', 'تاريخ العمارة والتصميم الداخلي 1',         'History of Architecture & Design 1', 3),
                ('INT105', 'تاريخ العمارة والتصميم الداخلي 2',         'History of Architecture & Design 2', 3),
                ('INT106', 'المنظور والظلال',                          'Perspective and Shadows',            4),
                ('INT107', 'خامات ومواد التصميم الداخلي',              'Interior Design Materials',          3),
                ('INT108', 'التصميم بمساعدة الحاسوب 2D',               'AutoCAD 2D',                         3),
                ('INT109', 'التصميم ثلاثي الأبعاد 3D Max',             'Computer 3D Modeling',               4),
                ('INT110', 'تكنولوجيا الإضاءة والصوتيات',              'Lighting and Acoustics Technology',  3),
            ],
        },
        {
            'name_ar': 'الهندسة', 'name_en': 'Engineering',
            'courses': [
                ('ENG101', 'رياضيات هندسية 1',                        'Engineering Mathematics 1',          3),
                ('ENG102', 'رياضيات هندسية 2',                        'Engineering Mathematics 2',          3),
                ('ENG103', 'فيزياء عامة وتطبيقية',                    'General and Applied Physics',        3),
                ('ENG104', 'كيمياء هندسية',                           'Engineering Chemistry',              3),
                ('ENG105', 'الرسم الهندسي بالحاسوب',                  'Computer Aided Engineering Drawing', 3),
                ('ENG106', 'ميكانيكا هندسية: استاتيكا',               'Engineering Mechanics: Statics',     3),
                ('ENG107', 'ميكانيكا هندسية: ديناميكا',               'Engineering Mechanics: Dynamics',    3),
                ('ENG108', 'مقدمة في البرمجة للهندسة',                'Introduction to Engineering Programming', 3),
                ('ENG109', 'مقاومة المواد الهندسية',                  'Strength of Engineering Materials',  3),
                ('ENG110', 'ميكانيكا الموائع',                        'Fluid Mechanics',                    3),
            ],
        },
        {
            'name_ar': 'المحاسبة', 'name_en': 'Accounting',
            'courses': [
                ('ACC101', 'مبادئ المحاسبة المالية 1',                'Principles of Financial Accounting 1', 3),
                ('ACC102', 'مبادئ المحاسبة المالية 2',                'Principles of Financial Accounting 2', 3),
                ('ACC103', 'المحاسبة في الشركات الشريكة',             'Accounting for Partnerships',          3),
                ('ACC104', 'المحاسبة في الشركات المساهمة',            'Accounting for Corporations',          3),
                ('ACC105', 'محاسبة التكاليف 1',                       'Cost Accounting 1',                    3),
                ('ACC106', 'محاسبة التكاليف 2',                       'Cost Accounting 2',                    3),
                ('ACC107', 'المحاسبة الإدارية',                       'Managerial Accounting',                3),
                ('ACC108', 'المحاسبة الحكومية والمنظمات غير الربحية', 'Governmental & Non-Profit Accounting', 3),
                ('ACC109', 'المحاسبة الضريبية والزكاة',               'Tax and Zakat Accounting',             3),
                ('ACC110', 'النظم المحاسبية الإلكترونية',             'Electronic Accounting Systems',        3),
            ],
        },
        {
            'name_ar': 'إدارة أعمال', 'name_en': 'Business Administration',
            'courses': [
                ('BUS101', 'مبادئ إدارة الأعمال 1',                   'Principles of Business Administration 1', 3),
                ('BUS102', 'مبادئ إدارة الأعمال 2',                   'Principles of Business Administration 2', 3),
                ('BUS103', 'إدارة الموارد البشرية',                   'Human Resource Management',               3),
                ('BUS104', 'السلوك التنظيمي في المؤسسات',             'Organizational Behavior',                 3),
                ('BUS105', 'إدارة التسويق الرقمي والتقليدي',          'Marketing Management',                    3),
                ('BUS106', 'الإدارة المالية والتمويل',                'Financial Management and Finance',        3),
                ('BUS107', 'إدارة الإنتاج والعمليات',                 'Production and Operations Management',    3),
                ('BUS108', 'إدارة المشاريع الصغيرة وريادة الأعمال',   'Small Business & Entrepreneurship',       3),
                ('BUS109', 'التخطيط والإدارة الإستراتيجية',           'Strategic Planning & Management',         3),
                ('BUS110', 'نظم المعلومات الإدارية MIS',              'Management Information Systems',          3),
            ],
        },
        {
            'name_ar': 'اللغات', 'name_en': 'Languages',
            'courses': [
                ('LNG101', 'مهارات القراءة والاستيعاب',               'Reading and Comprehension Skills',   3),
                ('LNG102', 'قواعد اللغة الأساسية 1',                  'Basic Language Grammar 1',           3),
                ('LNG103', 'قواعد اللغة المتقدمة 2',                  'Advanced Language Grammar 2',        3),
                ('LNG104', 'مهارات الكتابة والتعبير الأكاديمي',       'Academic Writing and Essay Skills',  3),
                ('LNG105', 'مهارات الاستماع والمحادثة',               'Listening and Speaking Skills',      3),
                ('LNG106', 'علم الصوتيات والفونيتكس',                 'Phonetics and Phonology',            3),
                ('LNG107', 'مقدمة في علم اللغويات',                   'Introduction to Linguistics',        3),
                ('LNG108', 'علم بناء الجملة والتركيب',                'Syntax and Sentence Structure',      3),
                ('LNG109', 'علم دلالات الألفاظ والمعاني',             'Semantics and Pragmatics',           3),
                ('LNG110', 'مبادئ الترجمة العامة',                    'Principles of General Translation',  3),
            ],
        },
        {
            'name_ar': 'الطب', 'name_en': 'Medicine',
            'courses': [
                ('MED101', 'علم التشريح البشري 1',                    'Human Anatomy 1',                    4),
                ('MED102', 'علم التشريح البشري 2',                    'Human Anatomy 2',                    4),
                ('MED103', 'علم وظائف الأعضاء الفسيولوجي 1',          'Medical Physiology 1',               4),
                ('MED104', 'علم وظائف الأعضاء الفسيولوجي 2',          'Medical Physiology 2',               4),
                ('MED105', 'الكيمياء الحيوية الطبية 1',               'Medical Biochemistry 1',             3),
                ('MED106', 'الكيمياء الحيوية الطبية والوراثية 2',     'Medical Biochemistry & Genetics 2',  3),
                ('MED107', 'علم الأنسجة والخلايا الهستولوجي',         'Histology and Cytology',             3),
                ('MED108', 'علم الأجنة والنمو الطبيعي',               'Embryology and Human Development',   3),
                ('MED109', 'علم الأمراض الباثولوجي العام 1',          'General Pathology 1',                4),
                ('MED110', 'علم الأمراض الخاص والجهازي 2',            'Systemic Pathology 2',               4),
            ],
        },
    ]

    for dept_data in SEED:
        dept = Department.query.filter_by(name_ar=dept_data['name_ar']).first()
        if not dept:
            dept = Department(name_ar=dept_data['name_ar'], name_en=dept_data['name_en'])
            db.session.add(dept)
            db.session.flush()          # get dept.id before inserting courses
        for code, name_ar, name_en, credits in dept_data['courses']:
            if not Course.query.filter_by(code=code).first():
                db.session.add(Course(
                    code=code, name_ar=name_ar, name_en=name_en,
                    credits=credits, department_id=dept.id,
                ))
    db.session.commit()
    print('✓ Seed: 7 departments / 70 courses verified.')

    # ═══════════════════════════════════════════════════════════════
    #  SEED ADVANCED — 7 Departments × 5 Advanced Courses (111–115)
    #  Prerequisites link back to introductory 101/102 courses.
    #  Department is derived automatically from the prereq's dept.
    # ═══════════════════════════════════════════════════════════════
    SEED_ADVANCED = [
        # Law
        ('LAW111', 'قانون العقود المقارن',                    'Comparative Contract Law',              3, 'LAW101'),
        ('LAW112', 'قانون الإجراءات المدنية',                 'Civil Procedure Law',                   3, 'LAW101'),
        ('LAW113', 'الأنظمة القانونية المقارنة',              'Comparative Legal Systems',             3, 'LAW102'),
        ('LAW114', 'قانون التجارة والشركات',                  'Commercial and Corporate Law',          3, 'LAW102'),
        ('LAW115', 'قانون حقوق الإنسان الدولي',               'International Human Rights Law',        3, 'LAW101'),
        # Interior Design
        ('INT111', 'تصميم المساكن الخاصة',                   'Residential Interior Design',           3, 'INT101'),
        ('INT112', 'تصميم الفضاءات التجارية',                 'Commercial Space Design',               3, 'INT101'),
        ('INT113', 'مشروع التصميم الداخلي 1',                 'Interior Design Project 1',             4, 'INT102'),
        ('INT114', 'برامج التصميم المتقدمة',                  'Advanced Design Software',              3, 'INT102'),
        ('INT115', 'التصميم المستدام والبيئي',                 'Sustainable and Environmental Design',  3, 'INT101'),
        # Engineering
        ('ENG111', 'التحليل الإنشائي الأساسي',                'Structural Analysis',                   3, 'ENG101'),
        ('ENG112', 'الهندسة الكهربائية الأساسية',              'Basic Electrical Engineering',          3, 'ENG101'),
        ('ENG113', 'هندسة البيئة والتلوث',                    'Environmental Engineering',             3, 'ENG102'),
        ('ENG114', 'المواد الهندسية المتقدمة',                 'Advanced Engineering Materials',        3, 'ENG102'),
        ('ENG115', 'ترموديناميكا هندسية',                     'Engineering Thermodynamics',            3, 'ENG101'),
        # Accounting
        ('ACC111', 'مراجعة الحسابات والتدقيق 1',              'Auditing and Assurance 1',              3, 'ACC101'),
        ('ACC112', 'مراجعة الحسابات والتدقيق 2',              'Auditing and Assurance 2',              3, 'ACC102'),
        ('ACC113', 'المحاسبة المتوسطة',                       'Intermediate Accounting',               3, 'ACC102'),
        ('ACC114', 'محاسبة القطاع المصرفي',                   'Banking Sector Accounting',             3, 'ACC101'),
        ('ACC115', 'تحليل القوائم المالية',                    'Financial Statement Analysis',          3, 'ACC101'),
        # Business Administration
        ('BUS111', 'إدارة سلسلة الإمداد والتوريد',            'Supply Chain Management',               3, 'BUS101'),
        ('BUS112', 'إدارة الجودة الشاملة TQM',                'Total Quality Management',              3, 'BUS101'),
        ('BUS113', 'الأعمال الدولية والعولمة',                 'International Business and Globalization', 3, 'BUS102'),
        ('BUS114', 'إدارة العلاقة مع العملاء CRM',             'Customer Relationship Management',      3, 'BUS102'),
        ('BUS115', 'قيادة الفرق وإدارة التغيير',               'Team Leadership and Change Management', 3, 'BUS101'),
        # Languages
        ('LNG111', 'تحليل النصوص الأدبية المقارن',             'Comparative Literary Text Analysis',    3, 'LNG101'),
        ('LNG112', 'الترجمة التخصصية التقنية والقانونية',      'Technical and Legal Translation',       3, 'LNG102'),
        ('LNG113', 'مهارات الكتابة الاحترافية',                'Professional Writing Skills',           3, 'LNG102'),
        ('LNG114', 'اللسانيات التطبيقية',                     'Applied Linguistics',                   3, 'LNG101'),
        ('LNG115', 'لغة إضافية: مستوى أول',                   'Second Language Level 1',               3, 'LNG101'),
        # Medicine
        ('MED111', 'مقدمة في علم الميكروبيولوجيا الطبية',     'Introduction to Medical Microbiology',  4, 'MED101'),
        ('MED112', 'علم الأدوية والصيدلة 1',                  'Pharmacology 1',                        3, 'MED101'),
        ('MED113', 'علم الأمراض الجزيئي والوراثي',            'Molecular and Genetic Pathology',       4, 'MED102'),
        ('MED114', 'مبادئ الفحص السريري',                     'Principles of Clinical Examination',    3, 'MED102'),
        ('MED115', 'المهارات الطبية السريرية الأساسية',         'Basic Clinical Medical Skills',         3, 'MED101'),
    ]

    for code, name_ar, name_en, credits, prereq_code in SEED_ADVANCED:
        if not Course.query.filter_by(code=code).first():
            prereq = Course.query.filter_by(code=prereq_code).first()
            if prereq:
                db.session.add(Course(
                    code=code, name_ar=name_ar, name_en=name_en,
                    credits=credits,
                    department_id=prereq.department_id,
                    prerequisite_id=prereq.id,
                ))
    db.session.commit()
    print('✓ Seed: 35 advanced courses (111–115 per dept) verified.')

    # ── Personal pre-reg entry (from Render env) ──────────────────
    if MY_STUDENT_ID and not PreRegisteredStudent.query.filter_by(student_id=MY_STUDENT_ID).first():
        db.session.add(PreRegisteredStudent(full_name=MY_FULL_NAME or MY_STUDENT_ID,
                                            student_id=MY_STUDENT_ID))
        db.session.commit()

    # ── Admin account (from Render env) ───────────────────────────
    if ADMIN_EMAIL:
        admin_user = User.query.filter_by(email=ADMIN_EMAIL).first()
        hashed_pw  = generate_password_hash(ADMIN_PASSWORD or 'ChangeMe123!')
        if not admin_user:
            admin_user = User(
                email=ADMIN_EMAIL, password_hash=hashed_pw,
                full_name='مدير النظام', student_id='ADMIN_2026',
                role='admin', status='approved',
            )
            db.session.add(admin_user)
        else:
            admin_user.password_hash = hashed_pw
        db.session.commit()

# تشغيل التطبيق (نهاية الملف)
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)

# --- نهاية الجزء الجاهز ---